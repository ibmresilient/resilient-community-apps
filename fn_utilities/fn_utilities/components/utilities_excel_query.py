# -*- coding: utf-8 -*-
# (c) Copyright IBM Corp. 2010, 2018. All Rights Reserved.
# pragma pylint: disable=unused-argument, no-self-use
"""
    Function utilities_excel_query receives attachment information and two additional fields -
    list of excel style ranges and comma separated list of names for defined ranges. It then
    gets the attachment and gets the data from it.

    It uses Openpyxl library (http://openpyxl.readthedocs.io/en/stable/).

    Logic for extracting data and parsing inputs is in the WorksheetData class.
"""

import logging
from resilient_circuits import ResilientComponent, function, handler, StatusMessage, FunctionResult, FunctionError
import openpyxl
import re  # to extract ranges from user input
import io  # to pass attachment to openpyxl
import datetime  # to deal with sheet's date fields
import json


class FunctionComponent(ResilientComponent):
    """Component that implements Resilient function 'utilities_excel_query"""

    def __init__(self, opts):
        """constructor provides access to the configuration options"""
        super(FunctionComponent, self).__init__(opts)
        self.options = opts.get("fn_utilities", {})

    @handler("reload")
    def _reload(self, event, opts):
        """Configuration options have changed, save new values"""
        self.options = opts.get("fn_utilities", {})

    @function("utilities_excel_query")
    def _utilities_excel_query_function(self, event, *args, **kwargs):
        """Function: Extract ranges of data or named ranges specified by the user from an excel document."""
        try:
            # Get the function parameters:
            # Attachment information
            yield StatusMessage("Starting the function")
            attachment_id = kwargs.get("attachment_id", None)  # number
            incident_id = kwargs.get("incident_id", None)  # number
            task_id = kwargs.get("task_id", None)  # number
            # What data to get
            excel_ranges = kwargs.get("excel_ranges", None)  # text
            excel_defined_names = kwargs.get("excel_defined_names", None)  # text

            log = logging.getLogger(__name__)
            log.info("task_id: {0}".format(task_id))
            log.info("incident_id: {0}".format(incident_id))
            log.info("attachment_id: {0}".format(attachment_id))
            log.info("excel_ranges: {0}".format(excel_ranges))
            log.info("excel_defined_names: {0}".format(excel_defined_names))

            if excel_defined_names is None:
                excel_defined_names = ""
            if excel_ranges is None:
                excel_ranges = ""

            # PUT YOUR FUNCTION IMPLEMENTATION CODE HERE
            yield StatusMessage("Reading the attachment.")
            attachment_data = self._get_attachment_binary(task_id, incident_id, attachment_id)
            yield StatusMessage("Processing attachment.")
            worksheet_data = WorksheetData(io.BytesIO(attachment_data), {
                "ranges": WorksheetData.parse_excel_notation(excel_ranges),
                "named_ranges": WorksheetData.parse_defined_names_notation(excel_defined_names)
            })
            worksheet_data.parse()  # extracts all the data to result
            worksheet_data.serialize()
            result = worksheet_data.result
            # Produce a FunctionResult with the results
            yield StatusMessage("Done.")
            yield FunctionResult(result)
        except Exception as e:
            log.error(str(e))
            yield FunctionError("A function error occurred.")

    def _get_attachment_binary(self, task_id, incident_id, attachment_id):
        """
        Gets the attachment with the rest client as binary.
        :param task_id: Number
            task id provided by user
        :param incident_id: Number
            incident id provided by user
        :param attachment_id: Number
            attachment id provided by user
        :return: Binary object
        """
        if incident_id is None and task_id is None:
            raise FunctionError("Either incident id or the task id has to be specified.")
        if attachment_id is None:
            raise FunctionError("The attachment id has to be specified.")
        if task_id:
            data_uri = "/tasks/{}/attachments/{}/contents".format(task_id, attachment_id)
        else:
            data_uri = "/incidents/{}/attachments/{}/contents".format(incident_id, attachment_id)

        client = self.rest_client()
        data = client.get_content(data_uri)  # type: object
        return data


class WorksheetData(object):
    """
    Facade pattern with openpyxl library that will parse the given worksheet into
    the json of requested content.
    """
    # This reg exp captures the form "Name"!A1:B2 or 'Name'!A1
    # Has 4 capturing groups - name of the sheet, top_left and bottom_right (if range given None if not), and cell
    EXCEL_RANGE_REG_EXP = r"(?:(?:\"|\')([\w\ ,;.\"\']+)(?:\"|\')!" \
                          r"(?:(?:([a-zA-Z]+[\d]+)[:]([a-zA-Z]+[\d]+))|([a-zA-Z]+[\d]+)))+"
    # These constant adjust the return JSON namings
    SHEETS_TITLES = "titles"
    PARSED_SHEETS = "sheets"
    PARSED_SHEETS_LIST = "parsed_sheets"
    NAMED_RANGES = "named_ranges"

    def __init__(self, path, opts={}):
        """
        :param path : String
            an absolute path to the file
        :param opts: Dict
            contains options with the following options:
            named_ranges: List of String names of the defined_names
            ranges: List of Objects that have name, top-left, and bottom-right
                coordinates of the range to be grabbed.
        """
        super(WorksheetData, self).__init__()
        self._file_path = path
        try:
            self.wb = openpyxl.load_workbook(self._file_path, read_only=True)
        except IOError as e:
            log = logging.getLogger(__name__)
            log.error(str(e))
            raise FunctionError("Error opening the provided file.")
        # options
        self.opts = opts
        # the eventual return value
        self.result = {}

    @staticmethod
    def serializer(obj):
        """
        when JSON passes a type it cannot serialize, check if it's one of the
        types that can come from a worksheet and serialize it.
        :return: String
            serialized field
        """
        if isinstance(obj, datetime.datetime):
            # the standard of Resilient is in microseconds
            return (obj - datetime.datetime(1970, 1, 1)).total_seconds()*1000
        return str(obj)

    @staticmethod
    def parse_excel_notation(ranges):
        """
        Takes in a string that has comma separated excel notation ranges
        :param ranges: String
            string
        :return: List[object]
            List of objects of a form {"name": "", "top_left": "", "bottom_right": ""}
        """
        if ranges is None:
            ranges = ""
        result = []
        range_matches = re.finditer(WorksheetData.EXCEL_RANGE_REG_EXP, ranges)
        for match in range_matches:
            name = match.group(1)
            # check if 2 coordinates were provided, or 1
            if match.group(2):
                top_left = match.group(2)
                bottom_right = match.group(3)
            # if the top left isn't provided, then a single coordinate in group 4 is given
            if not match.group(2):
                top_left = bottom_right = match.group(4)
            result.append({
                "name": name,
                "top_left":top_left,
                "bottom_right": bottom_right
            })
        return result

    @staticmethod
    def parse_defined_names_notation(defined_names):
        """
        :param defined_names: Is a list that gets passed to function, with comma separated
            names of defined ranges to be extracted
        :return: List[String]
            the list of split names
        """
        if defined_names is None:
            defined_names = ""
        split_names = [x.strip() for x in defined_names.split(",") if len(x.strip()) > 0]
        return split_names

    def parse(self):
        """
        Goes through the options and fills our self.result accordingly.
        """
        self.result = {self.SHEETS_TITLES: self.wb.sheetnames}

        # check if "named_ranges" is in the opts, and is not falsy
        if "named_ranges" in self.opts and self.opts["named_ranges"]:
            self.parse_named_ranges(self.opts["named_ranges"])

        if "ranges" in self.opts and self.opts["ranges"]:
            self.parse_sheet_ranges(self.opts["ranges"])

        self.add_keys()

    def serialize(self):
        self.result = json.loads(json.dumps(self.result, default=self.serializer))

    def parse_named_ranges(self, named_ranges):
        """
        Gets a list or a string of named ranges from options
        and calls parse_named_range for each of them
        """
        self.result[self.NAMED_RANGES] = {}
        # check if a list of named ranges is requested or a single named range
        if isinstance(named_ranges, list):
            for name in named_ranges:
                self.parse_named_range(name)
        elif isinstance(named_ranges, str):
            self.parse_named_range(named_ranges)
        elif isinstance(named_ranges, bool) and named_ranges:
            for range in self.wb.defined_names.definedName:  # defined_names is a list of DefinedName objects
                self.parse_named_range(range.name)

    def parse_named_range(self, name):
        """
        :param name: String
             name of the named range to add to the return.
             It contains multiple ranges itself.
        """
        result = {}
        # check if the named range is really in the workbook
        if name not in self.wb.defined_names:
            raise FunctionError("The defined range {0} doesn't exist".format(name))

        ranges = self.wb.defined_names[name]  # get the data of the named range
        destinations = ranges.destinations  # returns a generator of tuples (ws title, range)
        for sheet_name, rng in destinations:
            if sheet_name not in result:
                result[sheet_name] = {}
            # worksheet
            ws = self.wb[sheet_name]
            try:
                result[sheet_name][rng] = ([[cell.value for cell in row] for row in ws[rng]])
            except ValueError as e:
                log = logging.getLogger(__name__)
                log.error(str(e))
                raise FunctionError("Requested range {0} is not correct.".format(rng.coord))

        self.result[self.NAMED_RANGES][name] = result

    def parse_sheet_ranges(self, ranges):
        """
        Gets the list of ranges provided by user from the options and processes it.
        :param ranges: List[object] - list of range objects
        """
        if self.PARSED_SHEETS not in self.result:
            self.result[self.PARSED_SHEETS] = {}
        # rng is for range, since range is a builtin name
        for rng in ranges:
            self.parse_sheet_range(rng)

    def parse_sheet_range(self, range):
        """
        Parses a particular range - extract the data from the sheet.`
        :param range: object
            stores information about the range - sheet, top_left, bottom_right
        """
        # make sure the worksheet actually exists
        try:
            ws = self.wb[range["name"]]
        except KeyError as e:
            log = logging.getLogger(__name__)
            log.error(str(e))
            raise FunctionError("The sheet {} provided by user doesn't exist".format(range["name"]))

        # additional thing to do for read only sheets to make sure only necessary data is read
        ws.calculate_dimension(force=True)

        try:
            data = ws[range["top_left"]:range["bottom_right"]]
        except ValueError as e:
            log = logging.getLogger(__name__)
            log.error(str(e))
            raise FunctionError("The range coordinates {0},{1} provided by user are incorrect".
                      format(range["top_left"], range["bottom_right"]))

        if range["name"] not in self.result[self.PARSED_SHEETS]:
            self.result[self.PARSED_SHEETS][range["name"]] = {}
        result = [[cell.value for cell in row] for row in data]
        range_name = "{0}:{1}".format(range["top_left"], range["bottom_right"])
        self.result[self.PARSED_SHEETS][range["name"]][range_name] = result

    def add_keys(self):
        """
        Goes through self.result and adds _keys list to every dictionary as
        a workaround for dictionary iteration in the post processing.
        So adds _keys for all the sheets, and all the ranges in "sheets"
        and keys for all named_ranges, their sheets, and their ranges in "defined_names"
        """
        if self.PARSED_SHEETS in self.result:
            self.result[self.PARSED_SHEETS]["_keys"] = \
                [sheet for sheet in self.result[self.PARSED_SHEETS].keys() if sheet != "_keys"]
            for sheet_name, sheet in self.result[self.PARSED_SHEETS].items():
                # _keys is not a sheet so don't iterate over it, would need it even if the top line was moved down
                # in case _keys had already been defined before
                if sheet_name == "_keys":
                    continue
                sheet["_keys"] = [rng for rng in sheet.keys() if rng != "_keys"]
        if self.NAMED_RANGES in self.result:
            self.result[self.NAMED_RANGES]["_keys"] = \
                [rng for rng in self.result[self.NAMED_RANGES].keys() if rng != "_keys"]
            for name, named_range in self.result[self.NAMED_RANGES].items():
                # _keys is not a sheet so don't iterate over it, would need it even if the top line was moved down
                # in case _keys had already been defined before
                if name == "_keys":
                    continue
                named_range["_keys"] = [sheet for sheet in named_range.keys() if sheet != "_keys"]
                for sheet_name, sheet in named_range.items():
                    if sheet_name == "_keys":
                        continue
                    sheet["_keys"] = [rng for rng in sheet.keys() if rng != "_keys"]
