# -*- coding: utf-8 -*-
# pragma pylint: disable=unused-argument, no-self-use
"""Function implementation"""

import logging
import six
import ast
import json
import base64
import openpyxl
from io import BytesIO
from openpyxl.utils import column_index_from_string
from resilient_circuits import ResilientComponent, function, handler, StatusMessage, FunctionResult, FunctionError


class FunctionComponent(ResilientComponent):
    """Component that implements Resilient function 'spreadsheet_write"""

    def __init__(self, opts):
        """constructor provides access to the configuration options"""
        super(FunctionComponent, self).__init__(opts)
        self.options = opts.get("fn_utilities", {})

    @handler("reload")
    def _reload(self, event, opts):
        """Configuration options have changed, save new values"""
        self.options = opts.get("fn_utilities", {})

    def write_rows(self, worksheet, rows):
        """Write rows to worksheet"""
        if not isinstance(rows, list):
            raise ValueError("Rows not valid: {}".format(rows))
        rownum = 0
        for row in rows:
            rownum = rownum + 1
            if not isinstance(row, dict):
                raise ValueError("Row not valid: {}".format(row))
            for cell_column, cell_value in row.items():
                col_index = cell_column
                if isinstance(cell_column, six.string_types):
                    col_index = column_index_from_string(cell_column)
                worksheet.cell(row=rownum, column=col_index, value=cell_value)
        return

    @function("spreadsheet_write")
    def _spreadsheet_write_function(self, event, *args, **kwargs):
        """Function: Writes the data as a spreadsheet file attachment.

           The input is expected to be
               a list of rows, or a dictionary of {"sheet name": [list_of_rows]}, where
               each row is a dictionary of {"column_id": column_value}.

        """
        try:
            # Get the function parameters:
            input_data = kwargs.get("input_data")  # text

            log = logging.getLogger(__name__)
            log.info("input_data: %s", input_data)

            # The input data should be JSON,
            # alternatively it can be a Python str() representation
            data = None
            error = None
            try:
                data = json.loads(input_data)
            except ValueError:
                error = "Input data was not valid JSON."
                pass

            if data is None:
                try:
                    data = ast.literal_eval(input_data)
                except ValueError:
                    error = "Input data was not valid JSON."
                    pass

            if data is None:
                yield FunctionError(error or "Input data was not specified.")

            # Write the data into a spreadsheet
            workbook = openpyxl.Workbook()
            if isinstance(data, list):
                # Data is a list of rows
                sheet = workbook.active
                self.write_rows(sheet, data)
            elif isinstance(data, dict):
                # Data is a dictionary of {sheet name: list of rows}
                first = True
                for sheet_name, sheet_data in data.items():
                    if first:
                        sheet = workbook.active
                        sheet.title = sheet_name
                        first = False
                    else:
                        sheet = workbook.create_sheet(title=sheet_name)
                    self.write_rows(sheet, sheet_data)

            # Return base64-encoded spreadsheet data
            output = BytesIO()
            workbook.save(filename=output)
            b64value = base64.b64encode(output.getvalue())

            results = {
                "content": b64value
            }

            # Log the first chunk of the encoded result
            log.info("%s...%s, length=%d", b64value[:20], b64value[-20:], len(b64value))

            # Produce a FunctionResult with the return value
            yield FunctionResult(results)
        except Exception:
            yield FunctionError()
