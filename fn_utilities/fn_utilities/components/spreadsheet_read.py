# -*- coding: utf-8 -*-
# pragma pylint: disable=unused-argument, no-self-use
"""Function implementation"""

import logging
import json
import base64
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.read_only import EmptyCell
from io import BytesIO
from resilient_circuits import ResilientComponent, function, handler, StatusMessage, FunctionResult, FunctionError


class FunctionComponent(ResilientComponent):
    """Component that implements Resilient function 'spreadsheet_read"""

    @function("spreadsheet_read")
    def _spreadsheet_read_function(self, event, *args, **kwargs):
        """Function: Read data from a spreadsheet.

           The result will be "worksheets":
               a dictionary of {"sheet name": [list_of_rows]}, where
               each row is a dictionary of {"column_id": column_value}.

        """
        try:
            # Get the function parameters:
            base64content = kwargs.get("base64content")  # text
            single_sheet = kwargs.get("sheet")  # text
            range_string = kwargs.get("range")  # text, e.g. "A1:B5"
            cell_titles_row = 1  # Null means: cells titled A, B, etc.

            log = logging.getLogger(__name__)
            log.info("base64content: %s bytes", len(base64content))

            # Convert the input data to a byte stream
            binary_content = BytesIO(base64.b64decode(base64content))

            # Open as a spreadsheet
            xlsx = openpyxl.load_workbook(binary_content, read_only=True)

            # Enumerate all the sheets
            results = {"worksheets": {}}
            for sheet in xlsx.worksheets:
                rows = []
                results["worksheets"][sheet.title] = rows
                if single_sheet and sheet.title != single_sheet:
                    continue
                for row in sheet.iter_rows(range_string=range_string):
                    rowdata = {}
                    rows.append(rowdata)
                    for cell in row:
                        if not isinstance(cell, EmptyCell):
                            cellname = get_column_letter(cell.column)
                            if isinstance(cell.value, (datetime.datetime, datetime.date, datetime.time)):
                                rowdata[cellname] = cell.value.isoformat()
                            else:
                                rowdata[cellname] = cell.value

            # Produce a FunctionResult with the return value
            log.debug(json.dumps(results, indent=2))
            yield FunctionResult(results)
        except Exception:
            yield FunctionError()